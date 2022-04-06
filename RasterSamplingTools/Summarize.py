"""\
Summarize OutputDatabase

Created on September 15, 2021

@author Eric Mader
"""

import typing

import os
from sys import argv, stderr  #, stdout
import statistics
from openpyxl import Workbook
from openpyxl.worksheet import worksheet
from openpyxl.cell import cell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from TestArguments.CommandLineArguments import CommandLineOption, CommandLineArgs, ArgumentIterator
from UnicodeData.CharProps import getScript  #, scriptCodes
from UnicodeData.UCDTypeDictionaries import scriptNames as scriptCodes
from RasterSamplingTools.OutputDatabase import OutputDatabase

_usage = """
Usage:
summarize --input inputPath --output outputPath
"""

class SummarizeArgs(CommandLineArgs):
    widthFieldDict = {
        "mean": ["mean"],
        "median": ["median"],
        "most": ["min", "median", "mean", "max"],
        "all": ["min", "q1", "median", "mean", "q3", "max"],
    }

    @staticmethod
    def inFile(a: ArgumentIterator) -> str:
        return a.nextExtra("input file")


    options = [
        CommandLineOption("input", None, lambda a: a.nextExtra("input file"), "inputFile", None),
        CommandLineOption("output", None, lambda a: a.nextExtra("output file"), "outputFile", None),
    ]

    def __init__(self):
        CommandLineArgs.__init__(self)
        self._options.extend(SummarizeArgs.options)
        self.widthFields = self.widthFieldDict["most"]
        self.inputFile = ""
        self.outputFile = ""

def cellName(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"

def rangeFormula(row: int, minColumn: int, maxColumn: int) -> str:
    minCell = cellName(row, minColumn)
    maxCell = cellName(row, maxColumn)

    return f"={maxCell}-{minCell}"

def percentFormula(row: int, medianColumn: int, rangeColumn: int) -> str:
    medianCell = cellName(row, medianColumn)
    rangeCell = cellName(row, rangeColumn)

    return f"=IF({medianCell}<>0,{rangeCell}/{medianCell},\"\")"

def statCells(ws: worksheet.Worksheet, row: int, column: int, values: list[float], decimals: int = 1):
    numberFormat = f"0.{'0' * decimals}"
    for i, v in enumerate(values):
        vCell = typing.cast(cell.Cell, ws.cell(row=row, column=column + i, value=v))
        vCell.number_format = numberFormat
    
    rangeCell = typing.cast(cell.Cell,  ws.cell(row=row, column=column + 4, value=rangeFormula(row, column, column + 3)))
    rangeCell.number_format = numberFormat

    percentCell = typing.cast(cell.Cell,  ws.cell(row=row, column=column + 5, value=percentFormula(row, column + 1, column + 4)))
    percentCell.number_format = "0.0%"

def getScriptCode(codePoint: int) -> str:
    return scriptCodes[getScript(codePoint)]

def main():
    argumentList = argv
    args = None
    programName = os.path.basename(argumentList.pop(0))

    if len(argumentList) == 0:
        print(_usage, file=stderr)
        exit(1)

    try:
        args = SummarizeArgs()
        args.processArguments(argumentList)
        args.widthFields = SummarizeArgs.widthFieldDict["most"]  # need a better way to do this...
    except ValueError as error:
        print(programName + ": " + str(error), file=stderr)
        exit(1)

    # font = ctFont("Calibri", 11)

    outdb = OutputDatabase(args.inputFile)
    widthFields = args.widthFields
    fitResultFields = ["stroke_angle", "log_mean_orthogonal_distance"]

    wb = Workbook()
    ws = wb.active
    fieldNames = ["ps_name", "tested glyphs", "ignored glyphs", "scripts"]
    fieldNames.extend(widthFields)
    fieldNames.extend(["range", "range as % of median", "min angle", "median angle", "mean angle", "max angle", "angle range", "range as % of median", "min lmod", "median lmod", "mean lmod", "max lmod", "lmod range", "range as % of median"])

    for column, label in enumerate(fieldNames):
        labelCell = typing.cast(cell.Cell, ws.cell(row=1, column=column+1, value=label) )
        labelCell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rowNumber = 2
    # maxWidth = 0
    for entry in outdb.db:
        psName = entry["ps_name"]
        testResults = entry["test_results"]
        widths: dict[str, list[float]] = {wf: [] for wf in widthFields}
        fits: dict[str, list[float]] = {frf: [] for frf in fitResultFields}
        goodGlyphCount = 0
        haveWidths = False

        # maxWidth = max(maxWidth, stringWidth(psName, font))

        row = [psName]
        scripts: set[str] = set()

        for result in testResults.values():
            widthResults = result.get("widths", None)
            fitResults = result.get("fit_results", None)
            codePoints = result.get("code_points", None)

            for codePoint in codePoints:
                scripts.add(getScriptCode(codePoint))

            if widthResults:
                haveWidths = True
                goodGlyphCount += 1
                for widthField in widthFields:
                    widths[widthField].append(widthResults[widthField])

                for fitResultField in fitResultFields:
                    fits[fitResultField].append(fitResults[fitResultField])

        if haveWidths:
            means = [statistics.mean(widths[wf]) for wf in widthFields]

            angleFits = fits["stroke_angle"]
            angleMeans = [min(angleFits), statistics.median(angleFits), statistics.mean(angleFits), max(angleFits)]

            lmodFits = fits["log_mean_orthogonal_distance"]
            lmodMeans = [min(lmodFits), statistics.median(lmodFits), statistics.mean(lmodFits), max(lmodFits)]

            scripts.discard("Zzzz")  # Don't count the unknown script
            row.extend([goodGlyphCount, len(testResults) - goodGlyphCount, len(scripts)])

            ws.append(row)

            column = len(row) + 1
            statCells(ws, rowNumber, column, means)

            column += 6
            statCells(ws, rowNumber, column, angleMeans)

            column += 6
            statCells(ws, rowNumber, column, lmodMeans, decimals=4)

        else:
            ws.append([psName, goodGlyphCount, len(testResults) - goodGlyphCount])

        rowNumber += 1

    # ws.column_dimensions["A"].bestFit = True
    # ws.column_dimensions["A"].auto_size = True
    ws.column_dimensions["A"].width = 30  # type: ignore # maxWidth * .15?

    wb.save(args.outputFile)

if __name__ == "__main__":
    main()



